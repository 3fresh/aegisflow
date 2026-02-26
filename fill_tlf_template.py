"""
Fill TLF Template Script
Merge MOSAIC_CONVERT output Excel data with people_management data

This script will:
1. Prompt user to select MOSAIC_CONVERT output Excel file
2. Prompt user to select people_management.xlsx file
3. Merge MOSAIC data based on people_management file structure
4. Preserve all sheets and columns from people_management
5. Update merged data in corresponding sheet
6. Generate new Excel file and save to user-specified location
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import os
import tkinter as tk
from tkinter import filedialog
import sys
import time


def select_file(title, file_types):
    """
    Open file selection dialog
    
    Parameters:
    -----------
    title : str
        Dialog title
    file_types : list
        File types, e.g. [("Excel files", "*.xlsx"), ("All files", "*.*")]
    
    Returns:
    --------
    str : Selected file path, None if cancelled
    """
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=file_types
    )
    
    return file_path if file_path else None


def save_file(title, default_filename, file_types):
    """
    Open file save dialog
    
    Parameters:
    -----------
    title : str
        Dialog title
    default_filename : str
        Default filename
    file_types : list
        File types, e.g. [("Excel files", "*.xlsx"), ("All files", "*.*")]
    
    Returns:
    --------
    str : Selected save path, None if cancelled
    """
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.asksaveasfilename(
        title=title,
        defaultextension=".xlsx",
        initialfile=default_filename,
        filetypes=file_types
    )
    
    return file_path if file_path else None


def fill_tlf_template():
    """Main program: Merge MOSAIC data into people_management file"""
    
    print("=" * 80)
    print("TLF Template Filler - Fill TLF Template")
    print("=" * 80)
    
    # Step 1: Select input file
    print("\n[1] Please select MOSAIC_CONVERT output Excel file...")
    mosaic_file = select_file(
        "Select MOSAIC_CONVERT Output Excel File",
        [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not mosaic_file:
        print("❌ No MOSAIC_CONVERT file selected, program exiting")
        return False
    
    print(f"✓ MOSAIC_CONVERT file: {mosaic_file}")
    
    print("\n[2] Please select people_management.xlsx file...")
    people_file = select_file(
        "Select people_management.xlsx",
        [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not people_file:
        print("❌ No people_management file selected, program exiting")
        return False
    
    print(f"✓ People Management file: {people_file}")
    
    # Step 2: Read MOSAIC_CONVERT output
    print("\n[3] Reading MOSAIC_CONVERT data...")
    try:
        mosaic_df = None
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries and mosaic_df is None:
            try:
                mosaic_df = pd.read_excel(mosaic_file, sheet_name='Index', engine='openpyxl')
                print(f"✓ Read {len(mosaic_df)} rows of data")
                break
            except PermissionError:
                retry_count += 1
                if retry_count < max_retries:
                    print(f"  ⚠️ File is in use, retrying... ({retry_count}/{max_retries})")
                    time.sleep(1)
                else:
                    raise PermissionError(f"Unable to read MOSAIC file (file may be open in Excel):\n{mosaic_file}")
    except Exception as e:
        print(f"❌ Failed to read MOSAIC_CONVERT file: {e}")
        return False
    
    # Step 3: Read all sheets and structure from people_management file
    print("\n[4] Reading people_management file structure...")
    try:
        # Get all sheet information
        xls = pd.ExcelFile(people_file, engine='openpyxl')
        sheet_names = xls.sheet_names
        print(f"✓ Found {len(sheet_names)} sheets: {', '.join(sheet_names)}")
        
        # Read TLF sheet (for people data matching)
        target_sheet = 'TLF' if 'TLF' in sheet_names else sheet_names[0]
        print(f"✓ Will use '{target_sheet}' sheet for people data matching")
        
        # Read target sheet to get people data
        people_df = pd.read_excel(people_file, sheet_name=target_sheet, header=1, engine='openpyxl')
        print(f"  - '{target_sheet}' sheet: {len(people_df)} rows, {len(people_df.columns)} columns")
        
    except Exception as e:
        print(f"❌ Failed to read people_management file: {e}")
        return False
    
    # Step 4: Prepare MOSAIC data and perform three-tier matching
    print("\n[5] Processing data mapping and merging...")
    
    # Column mapping: source_column(MOSAIC) -> target_column(people_file)
    column_map = {
        'Output Type (Table, Listing, Figure)': 'Output Type (Table, Listing, Figure)',
        'tocnumber': 'Output # ',
        'Title': 'Title',
        'sect_num': 'Section # ',
        'sect_ttl': 'Section Title',
        'azsolid': 'Standard Template Reference',
        'PROGRAM': 'Program Name',
        'OUTFILE': 'Output Name'
    }
    
    # Build temporary dataframe for merging
    mosaic_merge_data = pd.DataFrame()
    for src_col, tgt_col in column_map.items():
        if src_col in mosaic_df.columns:
            mosaic_merge_data[tgt_col] = mosaic_df[src_col].values
        else:
            print(f"⚠️ Warning: Column {src_col} not found in source file")
            mosaic_merge_data[tgt_col] = None
    
    # Step 5: Merge people_management data - three-tier cascading match
    print("\n[6] Merging people data (three-tier cascading match)...")
    
    # Initialize tracking columns
    mosaic_merge_data['_tier1_matched'] = False
    mosaic_merge_data['_tier2_matched'] = False
    mosaic_merge_data['Programmer'] = None
    mosaic_merge_data['QC Program'] = None
    mosaic_merge_data['QC Programmer'] = None
    
    # ===== First priority: Match using Output Name =====
    if 'Output Name' in mosaic_merge_data.columns and 'Output Name' in people_df.columns:
        print("  [Step 1] Using Output Name for first priority matching...")
        
        people_by_outname = people_df[['Output Name', 'Programmer', 'QC Program', 'QC Programmer']].dropna(subset=['Output Name'])
        people_by_outname_unique = people_by_outname.drop_duplicates(subset=['Output Name'], keep='first')
        
        prog_lookup_outname = dict(zip(people_by_outname_unique['Output Name'], people_by_outname_unique['Programmer']))
        qc_prog_lookup_outname = dict(zip(people_by_outname_unique['Output Name'], people_by_outname_unique['QC Program']))
        qc_programmer_lookup_outname = dict(zip(people_by_outname_unique['Output Name'], people_by_outname_unique['QC Programmer']))
        
        for idx in mosaic_merge_data.index:
            outname = mosaic_merge_data.loc[idx, 'Output Name']
            if pd.notna(outname) and outname in prog_lookup_outname:
                mosaic_merge_data.loc[idx, 'Programmer'] = prog_lookup_outname[outname]
                mosaic_merge_data.loc[idx, 'QC Program'] = qc_prog_lookup_outname[outname]
                mosaic_merge_data.loc[idx, 'QC Programmer'] = qc_programmer_lookup_outname[outname]
                mosaic_merge_data.loc[idx, '_tier1_matched'] = True
        
        tier1_count = mosaic_merge_data['_tier1_matched'].sum()
        print(f"    ✓ Output Name matched - {tier1_count} rows")
    
    # ===== Second priority: Supplement match using Program Name =====
    if 'Program Name' in mosaic_merge_data.columns and 'Program Name' in people_df.columns:
        print("  [Step 2] Using Program Name for second priority matching (only supplement unmatched rows)...")
        
        people_by_progname = people_df[['Program Name', 'Programmer', 'QC Program', 'QC Programmer']].dropna(subset=['Program Name'])
        people_by_progname_unique = people_by_progname.drop_duplicates(subset=['Program Name'], keep='first')
        
        prog_lookup_progname = dict(zip(people_by_progname_unique['Program Name'], people_by_progname_unique['Programmer']))
        qc_prog_lookup_progname = dict(zip(people_by_progname_unique['Program Name'], people_by_progname_unique['QC Program']))
        qc_programmer_lookup_progname = dict(zip(people_by_progname_unique['Program Name'], people_by_progname_unique['QC Programmer']))
        
        for idx in mosaic_merge_data.index:
            if not mosaic_merge_data.loc[idx, '_tier1_matched']:
                progname = mosaic_merge_data.loc[idx, 'Program Name']
                if pd.notna(progname) and progname in prog_lookup_progname:
                    mosaic_merge_data.loc[idx, 'Programmer'] = prog_lookup_progname[progname]
                    mosaic_merge_data.loc[idx, 'QC Program'] = qc_prog_lookup_progname[progname]
                    mosaic_merge_data.loc[idx, 'QC Programmer'] = qc_programmer_lookup_progname[progname]
                    mosaic_merge_data.loc[idx, '_tier2_matched'] = True
        
        tier2_count = mosaic_merge_data['_tier2_matched'].sum()
        total_matched = mosaic_merge_data['_tier1_matched'].sum() + tier2_count
        print(f"    ✓ Program Name supplement match - {tier2_count} rows, total {total_matched} rows")
    
    # ===== Third step: Count unmatched rows =====
    unmatch_rows = mosaic_merge_data[~(mosaic_merge_data['_tier1_matched'] | mosaic_merge_data['_tier2_matched'])].index.tolist()
    green_highlight_rows = mosaic_merge_data[mosaic_merge_data['_tier2_matched']].index.tolist()
    
    if unmatch_rows:
        print(f"    ⚠️ Warning - Still {len(unmatch_rows)} rows unmatched (will be highlighted in yellow in output)")
    if green_highlight_rows:
        print(f"    ℹ️ Info - {len(green_highlight_rows)} rows matched via Program Name (will be highlighted in green in output)")
    
    # Step 6: Copy people_file workbook and update target sheet
    print("\n[7] Preparing output file...")
    
    try:
        # Load people_file workbook (preserve all sheets)
        wb = load_workbook(people_file)
        ws_target = wb[target_sheet]
        
        # Clear data from row 3 onwards (preserve first two header rows)
        print(f"  - Clearing data rows in '{target_sheet}' sheet (preserve headers)...")
        
        max_row = ws_target.max_row
        max_col = ws_target.max_column
        
        for row_idx in range(3, max_row + 1):
            for col_idx in range(1, max_col + 1):
                ws_target.cell(row=row_idx, column=col_idx).value = None
        
        # Get headers (row 2)
        headers = {}
        for col_idx in range(1, max_col + 1):
            header = ws_target.cell(row=2, column=col_idx).value
            if header:
                headers[header] = col_idx
        
        print(f"  - Found {len(headers)} column headers")
        
        # Write MOSAIC data to corresponding columns
        print(f"  - Filling {len(mosaic_merge_data)} rows of data...")
        
        for data_row_idx, (orig_idx, row_data) in enumerate(mosaic_merge_data.iterrows(), start=3):
            for col_name, col_idx in headers.items():
                if col_name in mosaic_merge_data.columns:
                    value = row_data[col_name]
                    if pd.isna(value):
                        value = None
                    ws_target.cell(row=data_row_idx, column=col_idx).value = value
        
        # Apply color highlighting
        print("  - Applying highlight colors...")
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        target_cols = ['Programmer', 'QC Program', 'QC Programmer']
        
        # Yellow highlight: Both Output Name and Program Name unmatched
        for row_idx in unmatch_rows:
            excel_row = row_idx + 3
            for col_name in target_cols:
                if col_name in headers:
                    col_idx = headers[col_name]
                    ws_target.cell(row=excel_row, column=col_idx).fill = yellow_fill
        
        # Green highlight: Matched via Program Name
        for row_idx in green_highlight_rows:
            excel_row = row_idx + 3
            for col_name in target_cols:
                if col_name in headers:
                    col_idx = headers[col_name]
                    ws_target.cell(row=excel_row, column=col_idx).fill = green_fill
        
        # Step 7: User selects save location
        print("\n[8] Please select output file save location...")
        output_file = save_file(
            "Save Output File",
            "people_management_updated.xlsx",
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_file:
            print("❌ No save location selected, program exiting")
            return False
        
        # Save file
        print(f"  - Saving file to: {output_file}")
        wb.save(output_file)
        print(f"✓ File saved: {output_file}")
        
    except Exception as e:
        print(f"❌ Failed to update Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    print("\n" + "=" * 80)
    print("✓✓✓ Filling complete!")
    print("=" * 80)
    print(f"Output file: {output_file}")
    print(f"Updated sheet: '{target_sheet}'")
    print(f"Data rows: {len(mosaic_merge_data)}")
    print(f"Output Name matched: {mosaic_merge_data['_tier1_matched'].sum()} rows")
    print(f"Program Name supplement match: {mosaic_merge_data['_tier2_matched'].sum()} rows")
    print(f"Unmatched (yellow highlight): {len(unmatch_rows)} rows")
    print(f"Tier 2 matched (green highlight): {len(green_highlight_rows)} rows")
    print("\nTip: You can directly open the Excel file to view results")
    print("      All sheets from people file have been preserved")
    
    return True


if __name__ == "__main__":
    try:
        success = fill_tlf_template()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ Program error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
