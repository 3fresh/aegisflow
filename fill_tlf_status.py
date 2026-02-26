"""
Fill TLF Status Script
Merge Comparison Status from TFL Status file to QC Status column in People Management file

This script will:
1. Prompt user to select the modified people_management.xlsx file
2. Prompt user to select the tfl_status.xlsx file
3. Preprocess Comparison Status in tfl_status (Match→Pass, Mismatch→Fail)
4. Perform exact matching based on Dataset and Output Name
5. Update QC Status column in TLF sheet of people_management
6. Preserve all other columns and sheets without any changes
7. Generate a new Excel file and save to user-specified location
8. Display statistics (total count, Pass count, Fail count, empty count)
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import os
import tkinter as tk
from tkinter import filedialog
import sys
import time


def select_file(title, file_types):
    """
    Open file selection dialog
    
    Parameters:
    -----------
    title : str
        Dialog title
    file_types : list
        File types, e.g. [("Excel files", "*.xlsx"), ("All files", "*.*")]
    
    Returns:
    --------
    str : Selected file path, or None if cancelled
    """
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=file_types
    )
    
    return file_path if file_path else None


def save_file(title, default_filename, file_types):
    """
    Open file save dialog
    
    Parameters:
    -----------
    title : str
        Dialog title
    default_filename : str
        Default filename
    file_types : list
        File types, e.g. [("Excel files", "*.xlsx"), ("All files", "*.*")]
    
    Returns:
    --------
    str : Selected save path, or None if cancelled
    """
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.asksaveasfilename(
        title=title,
        defaultextension=".xlsx",
        initialfile=default_filename,
        filetypes=file_types
    )
    
    return file_path if file_path else None


def fill_tlf_status():
    """Main program: Merge Comparison Status from TFL Status to QC Status column in People Management"""
    
    print("=" * 80)
    print("TLF Status Filler - Fill TLF Status")
    print("=" * 80)
    
    # Step 1: Select people_management file
    print("\n[1] Please select the modified people_management.xlsx file...")
    people_file = select_file(
        "Select people_management.xlsx file",
        [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not people_file:
        print("❌ No people_management file selected, exiting program")
        return False
    
    print(f"✓ People Management file: {people_file}")
    
    # Step 2: Select tfl_status file
    print("\n[2] Please select the tfl_status.xlsx file...")
    status_file = select_file(
        "Select tfl_status.xlsx file",
        [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not status_file:
        print("❌ No tfl_status file selected, exiting program")
        return False
    
    print(f"✓ TFL Status file: {status_file}")
    
    # Step 3: Read people_management file
    print("\n[3] Reading people_management file...")
    try:
        people_df = None
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries and people_df is None:
            try:
                # Read TLF sheet (header at row 2, index 1)
                # Use data_only=True to ignore formulas, only read values
                xl_people = pd.ExcelFile(people_file, engine='openpyxl')
                people_df = pd.read_excel(xl_people, sheet_name='TLF', header=1)
                xl_people.close()
                print(f"✓ Read {len(people_df)} rows of data")
                break
            except PermissionError:
                retry_count += 1
                if retry_count < max_retries:
                    print(f"  ⚠️ File is in use, retrying... ({retry_count}/{max_retries})")
                    time.sleep(1)
                else:
                    raise PermissionError(f"Cannot read people_management file (file may be open in Excel):\n{people_file}")
            except Exception as e:
                # If encountering date type error, try reading directly with openpyxl
                print(f"  ⚠️ Standard reading encountered issues, trying alternative method...")
                try:
                    # Use read_only=True and keep_vba=False to skip validation
                    wb_temp = load_workbook(people_file, read_only=True, keep_vba=False, data_only=True)
                    ws_temp = wb_temp['TLF']
                    
                    # Manually read data (row 2 is header, data starts from row 3)
                    headers_row = []
                    data_rows = []
                    
                    for idx, row in enumerate(ws_temp.iter_rows(values_only=True), start=1):
                        if idx == 2:
                            headers_row = list(row)
                        elif idx >= 3:
                            data_rows.append(row)
                    
                    people_df = pd.DataFrame(data_rows, columns=headers_row)
                    wb_temp.close()
                    print(f"✓ Read {len(people_df)} rows using alternative method")
                    break
                except Exception as e2:
                    print(f"❌ Alternative method also failed: {e2}")
                    raise e
    except Exception as e:
        print(f"❌ Failed to read people_management file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Step 4: Read tfl_status file
    print("\n[4] Reading tfl_status file...")
    try:
        status_df = None
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries and status_df is None:
            try:
                # Read Overview sheet
                xl_status = pd.ExcelFile(status_file, engine='openpyxl')
                status_df = pd.read_excel(xl_status, sheet_name='Overview')
                xl_status.close()
                print(f"✓ Read {len(status_df)} rows of status data")
                break
            except PermissionError:
                retry_count += 1
                if retry_count < max_retries:
                    print(f"  ⚠️ File is in use, retrying... ({retry_count}/{max_retries})")
                    time.sleep(1)
                else:
                    raise PermissionError(f"Cannot read tfl_status file (file may be open in Excel):\n{status_file}")
            except Exception as e:
                # If encountering date type error, try reading directly with openpyxl
                print(f"  ⚠️ Standard reading encountered issues, trying alternative method...")
                try:
                    # Use read_only=True and keep_vba=False to skip validation
                    wb_temp = load_workbook(status_file, read_only=True, keep_vba=False, data_only=True)
                    ws_temp = wb_temp['Overview']
                    
                    # Manually read data (row 1 is header)
                    headers_row = []
                    data_rows = []
                    
                    for idx, row in enumerate(ws_temp.iter_rows(values_only=True), start=1):
                        if idx == 1:
                            headers_row = list(row)
                        elif idx >= 2:
                            data_rows.append(row)
                    
                    status_df = pd.DataFrame(data_rows, columns=headers_row)
                    wb_temp.close()
                    print(f"✓ Read {len(status_df)} rows of status data using alternative method")
                    break
                except Exception as e2:
                    print(f"❌ Alternative method also failed: {e2}")
                    raise e
    except Exception as e:
        print(f"❌ Failed to read tfl_status file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Step 5: Preprocess Comparison Status (Match→Pass, Mismatch→Fail)
    print("\n[5] Preprocessing Comparison Status...")
    
    if 'Comparison Status' not in status_df.columns:
        print("❌ Error: 'Comparison Status' column not found in Overview sheet of tfl_status file")
        return False
    
    if 'Dataset' not in status_df.columns:
        print("❌ Error: 'Dataset' column not found in Overview sheet of tfl_status file")
        return False
    
    # Preprocessing: Match→Pass, Mismatch→Fail
    status_df['QC Status'] = status_df['Comparison Status'].replace({
        'Match': 'Pass',
        'Mismatch': 'Fail'
    })
    
    match_count = (status_df['Comparison Status'] == 'Match').sum()
    mismatch_count = (status_df['Comparison Status'] == 'Mismatch').sum()
    print(f"  ✓ Preprocessing complete: Match→Pass ({match_count} rows), Mismatch→Fail ({mismatch_count} rows)")
    
    # Step 6: Merge based on Dataset and Output Name
    print("\n[6] Merging QC Status data...")
    
    if 'Output Name' not in people_df.columns:
        print("❌ Error: 'Output Name' column not found in TLF sheet of people_management file")
        return False
    
    # Create Dataset→QC Status mapping dictionary
    status_lookup = dict(zip(status_df['Dataset'], status_df['QC Status']))
    
    print(f"  - Found {len(status_lookup)} status mappings")
    
    # Check if QC Status column exists, create if not
    qc_status_col = 'QC Status (Not Started, Ongoing, QC Pending, Fail, Pass)'
    if qc_status_col not in people_df.columns:
        people_df[qc_status_col] = None
        print(f"  - Created new column: {qc_status_col}")
    
    # Perform matching and merging
    matched_count = 0
    for idx in people_df.index:
        output_name = people_df.loc[idx, 'Output Name']
        if pd.notna(output_name) and output_name in status_lookup:
            people_df.loc[idx, qc_status_col] = status_lookup[output_name]
            matched_count += 1
        else:
            # Set unmatched rows to None
            people_df.loc[idx, qc_status_col] = None
    
    print(f"  ✓ Successfully matched {matched_count} rows")
    
    # Step 7: Calculate QC Status statistics
    print("\n[7] Calculating QC Status statistics...")
    
    total_count = len(people_df)
    pass_count = (people_df[qc_status_col] == 'Pass').sum()
    fail_count = (people_df[qc_status_col] == 'Fail').sum()
    empty_count = people_df[qc_status_col].isna().sum()
    
    print(f"  - Total TLF count: {total_count}")
    print(f"  - Pass count: {pass_count}")
    print(f"  - Fail count: {fail_count}")
    print(f"  - Empty count: {empty_count}")
    
    # Step 8: Update people_management file
    print("\n[8] Preparing output file...")
    
    try:
        # Load workbook (preserve all sheets and formats)
        wb = load_workbook(people_file)
        ws_tlf = wb['TLF']
        
        # Get headers (row 2)
        headers = {}
        max_col = ws_tlf.max_column
        
        for col_idx in range(1, max_col + 1):
            header = ws_tlf.cell(row=2, column=col_idx).value
            if header:
                headers[header] = col_idx
        
        print(f"  - Found {len(headers)} column headers")
        
        # Confirm QC Status column position
        if qc_status_col in headers:
            qc_col_idx = headers[qc_status_col]
            print(f"  - QC Status column is at column {qc_col_idx} (Excel column {get_column_letter(qc_col_idx)})")
        else:
            print(f"❌ Error: '{qc_status_col}' column not found in Excel")
            return False
        
        # Update QC Status column values
        print(f"  - Updating {len(people_df)} rows of QC Status data...")
        
        for data_row_idx, (_, row_data) in enumerate(people_df.iterrows(), start=3):
            value = row_data[qc_status_col]
            if pd.isna(value):
                value = None
            else:
                # Ensure value is string type (if not None)
                value = str(value) if value is not None else None
            ws_tlf.cell(row=data_row_idx, column=qc_col_idx).value = value
        
        # Step 9: User selects save location
        print("\n[9] Please select output file save location...")
        output_file = save_file(
            "Save output file",
            "people_management_with_status.xlsx",
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_file:
            print("❌ No save location selected, exiting program")
            return False
        
        # Save file
        print(f"  - Saving file to: {output_file}")
        wb.save(output_file)
        print(f"✓ File saved: {output_file}")
        
    except Exception as e:
        print(f"❌ Failed to update Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Step 10: Display final statistics
    print("\n" + "=" * 80)
    print("✓✓✓ Fill complete!")
    print("=" * 80)
    print(f"Output file: {output_file}")
    print(f"\nStatistics:")
    print(f"  - Total TLF count: {total_count}")
    print(f"  - Status 'Pass' count: {pass_count}")
    print(f"  - Status 'Fail' count: {fail_count}")
    print(f"  - Status empty count: {empty_count}")
    print(f"  - Match rate: {matched_count}/{total_count} ({matched_count/total_count*100:.1f}%)")
    print("\nTip: You can open the Excel file directly to view results")
    print("      All other columns and sheets have been preserved without any changes")
    
    return True


if __name__ == "__main__":
    try:
        success = fill_tlf_status()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ Program error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
