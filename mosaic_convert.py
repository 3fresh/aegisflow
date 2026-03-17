import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
import os
import tkinter as tk
from tkinter import filedialog
import re
from datetime import datetime


def preprocess_csv_data(df):
    """
    Preprocess CSV data: 
    1. Replace ''s with 's
    2. Replace ONLY first and last single quotes with double quotes
       in 'value' column when 'parm' starts with 'footnote' (NOT title)
    Middle quotes are kept unchanged as per requirement.
    
    Parameters:
    -----------
    df : pandas.DataFrame
        Input dataframe with 'parm' and 'value' columns
    
    Returns:
    --------
    df : pandas.DataFrame
        Preprocessed dataframe
    """
    def replace_quotes(value):
        """Replace ONLY first and last single quote with double quote"""
        if pd.isna(value) or value == '':
            return value
        
        value_str = str(value)
        
        # First, replace ''s with 's
        value_str = value_str.replace("''s", "'s")
        
        # Find first single quote
        first_quote_pos = value_str.find("'")
        # Find last single quote
        last_quote_pos = value_str.rfind("'")
        
        # Only process if there are single quotes
        if first_quote_pos != -1 and last_quote_pos != -1:
            if first_quote_pos == last_quote_pos:
                # Only one single quote, replace it
                value_str = value_str.replace("'", '"', 1)
            else:
                # Replace ONLY first and last single quotes with double quotes
                # Keep all middle quotes unchanged
                # Build new string: before_first + " + middle + " + after_last
                before_first = value_str[:first_quote_pos]
                middle = value_str[first_quote_pos + 1:last_quote_pos]
                after_last = value_str[last_quote_pos + 1:]
                value_str = before_first + '"' + middle + '"' + after_last
        
        return value_str
    
    # Resolve param column name (parm/param)
    if 'parm' in df.columns:
        param_col = 'parm'
    elif 'param' in df.columns:
        param_col = 'param'
    else:
        raise ValueError("Missing required column: 'parm' or 'param'")

    # Create a mask for rows where parm starts with 'footnote' (NOT title)
    mask = df[param_col].str.lower().str.startswith('footnote', na=False)
    
    # Apply the replacement only to matching rows
    df.loc[mask, 'value'] = df.loc[mask, 'value'].apply(replace_quotes)
    
    return df


def mosaic_convert(csv_file_path, output_file_path=None):
    """
    Convert TiFo CSV file to Excel format with MOSAIC processing
    
    Parameters:
    -----------
    csv_file_path : str
        Path to input CSV file
    output_file_path : str, optional
        Path to output XLSX file. If None, creates output in same directory with _MOSAIC_CONVERT suffix
    """
    
    # Read CSV file
    df = pd.read_csv(csv_file_path)

    # Resolve param column name (parm/param)
    if 'parm' in df.columns:
        param_col = 'parm'
    elif 'param' in df.columns:
        param_col = 'param'
    else:
        raise ValueError("Missing required column: 'parm' or 'param'")
    
    # Preprocess: Replace first and last single quotes in title/footnote values
    print("Preprocessing data...")
    df = preprocess_csv_data(df)
    
    # Get the number of rows
    lrow = len(df)

    # Step 0: Check tocnumber uniqueness
    toc_mask = df[param_col].str.lower().eq('tocnumber')
    toc_rows = df.loc[toc_mask, 'value']
    toc_row_count = int(toc_mask.sum())
    toc_unique_count = int(toc_rows.dropna().astype(str).str.strip().nunique())
    if toc_row_count > toc_unique_count:
        print("ERROR: Some tables/listings/figures use the same toc number")
        raise ValueError("ERROR: Some tables/listings/figures use the same toc number")
    
    # Step 1: Create progsfx column (G column) = program + suffix
    df['progsfx'] = df['program'] + df['suffix']
    
    # Step 2: Create columns H to AF with headers
    new_columns = ["sect_num", "sect_ttl", "outtype", "azsolid", "tocnumber", 
                   "Output Type (Table, Listing, Figure)", "Title", "PROGRAM", "SUFFIX", 
                   "OUTFILE", "title1", "title2", "title4", "title5", "title6", "title7", 
                   "footnote1", "footnote2", "footnote3", "footnote4", "footnote5", 
                   "footnote6", "footnote7", "footnote8", "footnote9"]
    
    # Step 3: No need to create formula columns on the original df
    # We'll do the pivot when creating the Index sheet
    
    # Step 4: Clean footnote7-like values that might be in 'value' column
    # This corresponds to the Replace operations in VBA
    df['value'] = df['value'].astype(str).replace("j=L '<<output program path>> <<output file name>> <<date/time>>' ", "", regex=False)
    df['value'] = df['value'].astype(str).replace("j=L '' ", "", regex=False)
    
    # Step 6: Create Index dataframe by pivoting the data using seq
    # 1. Use seq to convert multiple records per chart into a single row
    # 2. seq starts at 1 and increments by 1 each time param=outfile
    seq_values = []
    seq = 0
    for param_val in df[param_col].astype(str).fillna(''):
        if param_val.lower() == 'outfile':
            seq += 1
        seq_values.append(seq)
    df['seq'] = seq_values

    if seq == 0:
        raise ValueError("Missing required 'outfile' rows to build seq")

    index_data = []
    for seq_val, group in df.groupby('seq', sort=False):
        if seq_val == 0:
            continue
        group = group.copy()

        row_data = {
            'sect_num': group['sect_num'].iloc[0],
            'sect_ttl': group['sect_ttl'].iloc[0],
            'PROGRAM': group['program'].iloc[0] if pd.notna(group['program'].iloc[0]) else '',
            'SUFFIX': group['suffix'].iloc[0] if pd.notna(group['suffix'].iloc[0]) else '',
        }

        for _, row in group.iterrows():
            param_name = row.get(param_col, None)
            if pd.isna(param_name) or param_name == '':
                continue
            key = str(param_name).strip().lower()
            # Special handling: Convert 'outfile' to 'OUTFILE' (uppercase)
            if key == 'outfile':
                key = 'OUTFILE'
            row_data[key] = row.get('value', None)

        index_data.append(row_data)

    index_df = pd.DataFrame(index_data)
    
    # Step 7: OUTFILE is now directly from CSV (parm='outfile' -> value)
    # No need to build from PROGRAM + SUFFIX anymore
    
    # Set title1 to fixed value
    index_df['title1'] = "j=L 'AstraZeneca'"
    
    # Step 8: Extract Output Type from title4
    def determine_output_type(title_val):
        if pd.isna(title_val) or title_val == '' or title_val == 'None':
            return ''
        title_str = str(title_val)
        if 'Table' in title_str:
            return 'Table'
        elif 'Figure' in title_str:
            return 'Figure'
        elif 'Appendix' in title_str:
            return 'Listing'
        return ''
    
    if 'title4' in index_df.columns:
        index_df['Output Type (Table, Listing, Figure)'] = index_df['title4'].apply(determine_output_type)
    else:
        index_df['Output Type (Table, Listing, Figure)'] = ''
    
    # Step 9: Extract Title from title5
    def extract_title(title_val):
        if pd.isna(title_val) or title_val == '' or title_val == 'None':
            return ''
        title_str = str(title_val)
        # Normalize double-apostrophes (e.g. Investigator''s -> Investigator's)
        title_str = title_str.replace("''", "'")
        # Remove "j=C '" from start and "' " from end if present
        if title_str.startswith("j=C '") and title_str.endswith("' "):
            return title_str[5:-2]
        elif title_str.startswith("j=C '"):
            return title_str[5:]
        return title_str
    
    if 'title5' in index_df.columns:
        index_df['Title'] = index_df['title5'].apply(extract_title)
    else:
        index_df['Title'] = ''
    
    # Step 10: Clean up footnote columns - remove template placeholders
    footnote_cols = [col for col in index_df.columns if col.startswith('footnote')]
    for col in footnote_cols:
        if col in index_df.columns:
            # Remove template placeholder footnotes
            index_df[col] = index_df[col].apply(
                lambda x: '' if pd.notna(x) and 
                (str(x).strip() in ['j=L "<<output program path>> <<output file name>> <<date/time>>"',
                                    'j=L "<<output program path>> <<output file name>> <<date/time>>" ',
                                    "j=L '<<output program path>> <<output file name>> <<date/time>>'",
                                    "j=L '<<output program path>> <<output file name>> <<date/time>>' "]) 
                else x
            )
    
    # Step 10b: Convert title7 j=C to j=L if not empty
    if 'title7' in index_df.columns:
        def convert_title7(value):
            if pd.isna(value) or str(value).strip() == '':
                return value
            value_str = str(value)
            # Replace j=C with j=L
            value_str = value_str.replace("j=C '", "j=L '")
            return value_str
        index_df['title7'] = index_df['title7'].apply(convert_title7)
    
    # Step 10c: Normalize title5 - convert '' to ', then re-wrap with j=C '' ... '' delimiters
    if 'title5' in index_df.columns:
        def normalize_title5(value):
            if pd.isna(value) or str(value).strip() == '':
                return value
            value_str = str(value)
            # Normalize double-apostrophes first
            value_str = value_str.replace("''", "'")
            # Re-wrap SAS delimiters: j=C '...' (with or without trailing space) -> j=C ''...''
            if value_str.startswith("j=C '") and value_str.endswith("' "):
                inner = value_str[5:-2]
                return f"j=C ''{inner}''"
            elif value_str.startswith("j=C '"):
                inner = value_str[5:]
                if inner.endswith("'"):
                    inner = inner[:-1]
                return f"j=C ''{inner}''"
            return value_str
        index_df['title5'] = index_df['title5'].apply(normalize_title5)
    
    # Step 11: Reorder columns to match expected output
    final_columns = ['sect_num', 'sect_ttl', 'outtype', 'azsolid', 'tocnumber',
                     'Output Type (Table, Listing, Figure)', 'Title', 'PROGRAM', 'SUFFIX',
                     'OUTFILE', 'title1', 'title2', 'title4', 'title5', 'title6', 'title7',
                     'footnote1', 'footnote2', 'footnote3', 'footnote4', 'footnote5',
                     'footnote6', 'footnote7', 'footnote8', 'footnote9']
    
    # Ensure all columns exist
    for col in final_columns:
        if col not in index_df.columns:
            index_df[col] = None
    
    # Keep only the final columns in order
    index_final = index_df[final_columns].copy()
    
    # Insert "Core" column at position 4 (after azsolid, before tocnumber)
    index_final.insert(4, 'Core', '')
    
    # Sort by tocnumber using numeric comparison (not string comparison)
    # This ensures 14.1.2 comes before 14.1.10
    if 'tocnumber' in index_final.columns and index_final['tocnumber'].notna().any():
        def tocnumber_sort_key(tocnum):
            """Convert tocnumber to sortable tuple of integers for proper numeric sorting"""
            if pd.isna(tocnum):
                return (float('inf'),)  # Put NaN at the end
            try:
                # Split by dots and convert each part to int
                parts = str(tocnum).split('.')
                return tuple(int(p) if p.isdigit() else float('inf') for p in parts)
            except (ValueError, AttributeError):
                return (float('inf'),)
        
        # Create temporary sort key column
        index_final['_tocnumber_sort_key'] = index_final['tocnumber'].apply(tocnumber_sort_key)
        # Sort by the numeric key
        index_final = index_final.sort_values('_tocnumber_sort_key').drop('_tocnumber_sort_key', axis=1).reset_index(drop=True)
    
    # Step 11: Save to Excel with formatting
    if output_file_path is None:
        base_name = os.path.splitext(csv_file_path)[0]
        date_suffix = datetime.now().strftime('%Y%m%d')
        output_file_path = f"{base_name}_MOSAIC_CONVERT_{date_suffix}.xlsx"
    
    # Write to Excel
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        # Write the Index sheet
        index_final.to_excel(writer, sheet_name='Index', index=False)
        
        # Write the original data (cleaned)
        df_output = df[['sect_num', 'sect_ttl', 'program', 'suffix', param_col, 'value']].copy()
        df_output.to_excel(writer, sheet_name='Original', index=False)
    
    # Step 12: Apply formatting
    wb = load_workbook(output_file_path)
    ws = wb['Index']
    
    # Helper function to find all characters that need red highlighting
    def find_highlight_positions(text):
        """Find positions of characters that need to be highlighted in red (non-latin1 and ''s)"""
        if pd.isna(text) or text == '':
            return []
        if not isinstance(text, str):
            return []
        
        positions = set()
        
        # Find non-latin1 characters
        for i, char in enumerate(text):
            try:
                char.encode('latin1')
            except UnicodeEncodeError:
                positions.add(i)
        
        # Find ''s patterns (two single quotes followed by s)
        apostrophe_s_pattern = "''s"
        start = 0
        while True:
            pos = text.find(apostrophe_s_pattern, start)
            if pos == -1:
                break
            positions.add(pos)      # first '
            positions.add(pos + 1)  # second '
            positions.add(pos + 2)  # s
            start = pos + 1
        
        return sorted(list(positions))
    
    def create_rich_text(text, highlight_positions):
        """Create rich text with specified positions highlighted in red"""
        if not highlight_positions:
            return text

        if not isinstance(text, str):
            return text
        text_str = text
        rich_text_parts = []
        current_pos = 0
        
        # Define fonts
        default_inline = InlineFont(rFont='DengXian')
        red_inline = InlineFont(rFont='DengXian', color='FF0000')
        
        for pos in highlight_positions:
            # Add text before the highlighted char
            if pos > current_pos:
                rich_text_parts.append(TextBlock(default_inline, text_str[current_pos:pos]))
            # Add the highlighted char in red
            rich_text_parts.append(TextBlock(red_inline, text_str[pos]))
            current_pos = pos + 1
        
        # Add remaining text
        if current_pos < len(text_str):
            rich_text_parts.append(TextBlock(default_inline, text_str[current_pos:]))
        
        return CellRichText(*rich_text_parts)
    
    # Identify duplicate program+suffix combinations
    prog_suff_keys = index_final[['PROGRAM', 'SUFFIX']].fillna('').astype(str)
    prog_suff_key_series = prog_suff_keys['PROGRAM'] + '||' + prog_suff_keys['SUFFIX']
    dup_mask = prog_suff_key_series.duplicated(keep=False)
    dup_mask = dup_mask & (prog_suff_key_series != '||')
    dup_rows = set(index_final.index[dup_mask])

    if dup_rows:
        print("ERROR: Some charts share the same program+suffix, please update MOSAIC")
    
    # Find column indices for tocnumber, PROGRAM, SUFFIX
    col_indices = {}
    for idx, col_name in enumerate(index_final.columns, start=1):
        col_indices[col_name] = idx
    
    tocnum_col = col_indices.get('tocnumber')
    program_col = col_indices.get('PROGRAM')
    suffix_col = col_indices.get('SUFFIX')
    footnote_cols = {col_indices[col] for col in index_final.columns if col.startswith('footnote')}
    
    # Identify footnote columns with gaps (empty cells before last non-empty footnote)
    footnote_col_names = [col for col in index_final.columns if col.startswith('footnote')]
    empty_footnote_cells = set()  # Set of (row_idx, col_idx) tuples for Excel coordinates
    
    for data_row_idx in range(len(index_final)):
        # Get values for all footnote columns in this row
        footnote_info = []
        for col_name in footnote_col_names:
            col_idx = col_indices[col_name]
            value = index_final.loc[data_row_idx, col_name]
            is_empty = pd.isna(value) or str(value).strip() == ''
            footnote_info.append((col_idx, is_empty))
        
        # Find the last non-empty footnote column
        last_non_empty_idx = -1
        for i in range(len(footnote_info) - 1, -1, -1):
            col_idx, is_empty = footnote_info[i]
            if not is_empty:
                last_non_empty_idx = i
                break
        
        # Mark all empty footnote columns before the last non-empty one
        if last_non_empty_idx > 0:
            for i in range(last_non_empty_idx):
                col_idx, is_empty = footnote_info[i]
                if is_empty:
                    # Excel row = data_row_idx + 2 (header is row 1, data starts at row 2)
                    empty_footnote_cells.add((data_row_idx + 2, col_idx))
    
    # Define fill colors and fonts
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # Light green
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    default_font = Font(name='DengXian')
    bold_font = Font(name='DengXian', bold=True)
    
    # Format all cells
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        for col_idx, cell in enumerate(row, start=1):
            # Row 1: Headers - bold font
            if row_idx == 1:
                cell.font = bold_font
                # Highlight H1:J1 (columns 8, 9, 10)
                if 8 <= col_idx <= 10:
                    cell.fill = yellow_fill
            else:
                # Data rows
                cell_value = cell.value
                
                # Check if this row has duplicate program+suffix
                data_row_idx = row_idx - 2  # Convert to 0-based index for index_final
                is_dup_prog_suff = data_row_idx in dup_rows
                
                # Find all characters that need red highlighting (non-latin1 and ''s)
                text_value = '' if cell_value is None else str(cell_value)
                highlight_positions = find_highlight_positions(text_value)
                has_highlight_chars = len(highlight_positions) > 0

                # Check if this is an empty footnote cell with a gap
                is_empty_footnote_gap = (row_idx, col_idx) in empty_footnote_cells

                # Footnote cells must end with a double quote
                needs_quote_highlight = (
                    col_idx in footnote_cols and text_value.strip() != '' and not text_value.rstrip().endswith('"')
                )
                
                # Apply formatting based on conditions
                if has_highlight_chars:
                    # Has special chars: create rich text with those chars in red, green background
                    cell.value = create_rich_text(text_value, highlight_positions)
                    cell.fill = green_fill
                elif is_empty_footnote_gap:
                    # Empty footnote with gap: green background
                    cell.fill = green_fill
                    cell.font = default_font
                elif is_dup_prog_suff and col_idx in [program_col, suffix_col]:
                    # Duplicate PROGRAM+SUFFIX: yellow fill
                    cell.fill = yellow_fill
                    cell.font = default_font
                else:
                    # Default: DengXian font
                    cell.font = default_font

                if needs_quote_highlight:
                    cell.fill = blue_fill
    
    # Freeze panes at H2
    ws.freeze_panes = 'H2'
    
    wb.save(output_file_path)
    
    print(f"OK: Conversion complete! Output file: {output_file_path}")
    print(f"  - Processed {len(df)} rows of raw data")
    print(f"  - Generated {len(index_final)} rows of index data")
    
    return output_file_path


def select_input_file():
    """
    Open file dialog to select input CSV file
    
    Returns:
    --------
    str : Selected file path or None if cancelled
    """
    root = tk.Tk()
    root.withdraw()  # Hide main window
    root.attributes('-topmost', True)  # Bring dialog to front
    
    file_path = filedialog.askopenfilename(
        title="Select input CSV file",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        initialdir=os.getcwd()
    )
    
    root.destroy()
    return file_path if file_path else None


def select_output_file(default_name="output_MOSAIC_CONVERT.xlsx"):
    """
    Open file dialog to select output XLSX file
    
    Parameters:
    -----------
    default_name : str
        Default filename
    
    Returns:
    --------
    str : Selected file path or None if cancelled
    """
    root = tk.Tk()
    root.withdraw()  # Hide main window
    root.attributes('-topmost', True)  # Bring dialog to front
    
    file_path = filedialog.asksaveasfilename(
        title="Select output XLSX file",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name,
        initialdir=os.getcwd()
    )
    
    root.destroy()
    return file_path if file_path else None


if __name__ == "__main__":
    print("=" * 80)
    print("MOSAIC_CONVERT - CSV to Excel converter")
    print("=" * 80)
    print()
    
    # Step 1: Select input file
    print("[1] Please select input CSV file...")
    input_file = select_input_file()
    
    if not input_file:
        print("\nERROR: No input file selected, exiting.")
        exit(0)
    
    print(f"OK: Selected input file: {input_file}")
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"\nERROR: File not found: {input_file}")
        exit(1)
    
    # Step 2: Select output file
    print("\n[2] Please select output XLSX file...")
    
    # Suggest default output name based on input file
    input_base = os.path.splitext(os.path.basename(input_file))[0]
    date_suffix = datetime.now().strftime('%Y%m%d')
    default_output = f"{input_base}_MOSAIC_CONVERT_{date_suffix}"
    
    output_file = select_output_file(default_output)
    
    if not output_file:
        print("\nERROR: No output file selected, exiting.")
        exit(0)
    
    print(f"OK: Selected output file: {output_file}")
    
    # Step 3: Run the conversion
    print("\n[3] Starting conversion...")
    print("=" * 80)
    
    try:
        result_file = mosaic_convert(input_file, output_file)
        print("\n" + "=" * 80)
        print("OK: Conversion successful!")
        print("=" * 80)
        print(f"Output file: {result_file}")
        print("\nTip: You can open the Excel file directly to view the results")
        
        # Save output path to temp file for use by validation script (UTF-8 encoding for path support)
        with open(".last_output.txt", "w", encoding="utf-8") as f:
            f.write(result_file)
        
    except Exception as e:
        print("\n" + "=" * 80)
        print("ERROR: Conversion failed!")
        print("=" * 80)
        print(f"Error details: {e}")
        print("\nFull traceback:")
        import traceback
        traceback.print_exc()
        exit(1)
